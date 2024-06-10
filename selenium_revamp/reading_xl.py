from openpyxl import Workbook, load_workbook

wb = load_workbook(filename='demoexcel.xlsx')
# sh = wb.active  # to select the active sheet
# setting via sheet name:
sh = wb['DemoSheet']

# print(sh['A3'].value)
#
# print(sh.cell(2, 2).value)
row_count = sh.max_row
col_count = sh.max_column

print(row_count)
print(col_count)

for i in range(1, row_count+1):
    for j in range(1, col_count+1):
        print(sh.cell(row=i, column=j).value, end=" ")
    print()


# alternate way
for i in range(1, row_count+1):  # for every row
    print(tuple(i.value for i in sh[i]))  # for every col in a given row 
    