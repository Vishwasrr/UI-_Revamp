from faker import Faker
from openpyxl import Workbook
from random import randint

# fake_data = Faker(['en_US', 'ja_JP'])
fake_data = Faker()

wb = Workbook()
ws = wb.active

for i in range(1, 11):
    for j in range(1, 4):
        name = fake_data.name()
        names = name.split(" ")  # making sure the email is generated from name
        fname, lname = names[0].lower(), names[-1].lower()
        ws.cell(row=i, column=1).value = name
        ws.cell(row=i, column=2).value = fname + lname + str(randint(1, 100)) + "@gmail.com"
        ws.cell(row=i, column=3).value = fake_data.address()

wb.save("test_data.xlsx")
"""
This is a test docstring
"This is for double quotes"

"""
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())
